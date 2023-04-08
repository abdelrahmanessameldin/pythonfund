import openpyxl
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
#print (product_list.max_row)
#print (inv_file)

products_per_supplier = {}
total_value_per_supplier ={}
products_under_10 = {}

for product_row in range (2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    print (inventory)
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    #print (supplier_name)

    if supplier_name in products_per_supplier:

        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name]= current_num_products + 1

    else:

        print ("adding a new supplier")
        products_per_supplier[supplier_name] = 1

#calculation total valueo of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)    
        print (total_value_per_supplier.get(supplier_name) )
        total_value_per_supplier[supplier_name] = current_total_value + (price * inventory)
    else: 

        total_value_per_supplier[supplier_name] =   price * inventory

#logic_products_with_nventrory_less_10

    if inventory < 10 :
        products_under_10[product_number] = inventory

# add value for total ivnetory proce

    inventory_price.value =inventory * price

inv_file.save("inventory_with_total_value.xlsx")
inv_file.close()

print (products_per_supplier)
print (total_value_per_supplier)
print (products_under_10)
